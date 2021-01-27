import openpyxl


class  HomePageData():

    test_home_page_data = [
        {"first_name": "Hohoh", "email": "hoho@mail.com", "password": "1245AqwdfrrF!", "gender": "Male", "dofb": "10.11.1991"},
        {"first_name": "Aresw", "email": "werqwe.com", "password": "46q5we4", "gender": "Female", "dofb": "7.06.1999"},
        {"first_name": "Tweeq", "email": "aew@mail.com", "password": "srg89ewe", "gender": "Male", "dofb": "07.5.1971"}
    ]

    @staticmethod
    def get_test_data_from_excel(test_case_name):
        dict = {}

        book = openpyxl.load_workbook(filename="/Users/marko/Desktop/PythonTest/qa_automation_frwrk/TestData/PythonDemo.xlsx", data_only=True)
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict]
