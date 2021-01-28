import openpyxl
from utilities.BaseClass import BaseClass


dict_data = {}
log_info = BaseClass()
log = log_info.get_logger()

book = openpyxl.load_workbook(filename="/Users/marko/Desktop/PythonTest/qa_automation_frwrk/TestData/PythonDemo.xlsx", data_only=True)
sheet = book.active
cell = sheet.cell(row=2, column=2)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TestCase1":
        for j in range(2, sheet.max_column + 1):
            dict_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

log.info(dict_data)
